# -*- coding: utf-8 -*-
"""
Language:   python3
Goal:       Parse Face Recognization results
"""
import os
import sys
import time
from concurrent.futures import ProcessPoolExecutor, wait
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import getopt

### Defs region
def printUsage():
    print(u"""
Usage:
-h/--help:      show this help message
-i/--xlsFileR:  index excel file (.xlsx) path
-o/--xlsFileW:  output excel file (.xlsx) path
-d/--srcRoot:   directory where face recognition log folders are
-l/--checkList: a list of folder names in [srcRoot] seperated by '/'; if not
                specified, all direct sub folders in [srcRoot] are processed

E.g.:
$>python [this.script] -i ./facerecog-index.xlsx -d ./facerecog.log
  -o ./facerecog-interim-result.xlsx -l yawen/xingliujian/yukeke
""")

def parseOpt():
    try:
        opts, args = getopt.getopt(sys.argv[1:], 'hi:d:o:l:', \
                ['help', 'xlsFileR=', 'srcRoot=', 'xlsFileW=', 'checkList='])
    except getopt.GetoptError as err:
        print(str(err))
        sys.exit(1)
    for o, v in opts:
        if o in ('-h', '--help'):
            printUsage()
            sys.exit(1)
        elif o in ('-i', '--xlsFileR'):
            xlsFileR = v
        elif o in ('-o', '--xlsFileW'):
            xlsFileW = v
        elif o in ('-d', '--srcRoot'):
            srcRoot = v
        elif o in ('-l', '--checkList'):
            checkList = v.split("/")
    for o in ('xlsFileR', 'srcRoot', 'xlsFileW'):
        if not o in dir():
            print("Error: option not specified: "+o)
            printUsage()
            sys.exit(1)
    for o in ('xlsFileR', 'srcRoot'):
        v = locals()[o]
        if not os.path.exists(v):
            print("Error: path does not exist: "+v)
            printUsage()
            sys.exit(1)
    if not 'checkList' in dir():
        checkList=os.listdir(srcRoot)
    if not checkList: # empty checkList
        print("Error: empty checkList")
        printUsage()
        sys.exit(1)
    for v in checkList:
        if not os.path.exists(srcRoot+"/"+v):
            print("Error: folder does not exist: "+v)
            printUsage()
            sys.exit(1)
    return (xlsFileR, xlsFileW, srcRoot, checkList)

class RecgInfo:
    # __slots__ = ("startRecg", "endRecg", "endParse", "recgCost", "recgStatus",
    #     "score", "parseCost")
    def __init__(self):
        self.startRecg = None
        self.endRecg = None
        self.endParse = None
        self.recgCost = None

        self.recgStatus = None
        self.score = None
        self.parseCost = None
        self.recgIsValid = None

    def printResult(self):
        if (None != self.startRecg):
            print("startRecgTime:\t" + str(self.startRecg))
        else:
            print("startRecgTime:\tnull")
        if (None != self.endRecg):
            print("endRecgTime:\t" + str(self.endRecg))
        else:
            print("endRecgTime:\tnull")
        if (None != self.endParse):
            print("endParseTime:\t" + str(self.endParse))
        else:
            print("endParseTime:\tnull")
        if (None != self.startRecg and None != self.endRecg):
            self.recgCost = self.endRecg - self.startRecg
            print("recgCost:\t\t" + str(self.recgCost))
        else:
            print("recgCost:\t\tnull")

        if (None != self.endParse and None != self.endRecg):
            self.parseCost = self.endParse - self.endRecg
            print("parseCost:\t\t" + str(self.parseCost))
        else:
            print("parseCost:\t\tnull")
        if (None != self.recgStatus):
            print("recgStatus:\t\t" + self.recgStatus)
        else:
            print("recgStatus:\t\tnull")
        if (None != self.score):
            print("score:\t\t\t" + str(self.score))
        else:
            print("score:\t\t\tnull")

class LogInfo:
    # __slots__ = ("frameID", "caseID", "loadCost", "loadFile", "detCost",
    #     "detNum")
    def __init__(self, mFrameID, iFileID, iFileSN):
        self.frameID = mFrameID
        self.fileID = iFileID
        self.fileSN = iFileSN
        self.caseID = None
        self.__startTime = None
        self.__endTime = None
        self.__endStatus = None
        self.totalCost = None

        self.__startLoad = None
        self.__endLoad = None
        self.loadCost = None
        self.loadFile = None

        self.__startDet = None
        self.__endDet = None
        self.detCost = None
        self.detNum = None

        self.__recg = RecgInfo()

    def setST(self, startTime):
        self.__startTime = startTime
        self.__startLoad = startTime

    def setET(self, endTime, mEndStatus):
        if (None == self.__endTime):
            self.__endTime = endTime
            self.__endStatus = mEndStatus
        elif (endTime > self.__endTime):
            self.__endTime = endTime
            self.__endStatus = mEndStatus

    def setRecgValid(self, fIsValid):
        self.__recg.recgIsValid = fIsValid

    def setEL(self, endLoad, mCaseID, fileName):
        self.__endLoad = endLoad
        self.caseID = mCaseID
        self.loadFile = fileName

    def setSD(self, startDet):
        self.__startDet = startDet

    def setED(self, endDet, mDetNum):
        self.__endDet = endDet
        self.detNum = mDetNum

    def setSR(self, startRecg):
        if (None == self.__recg.startRecg):
            self.__recg.startRecg = startRecg

    def setER(self, endRecg):
        if (None == self.__recg.endRecg):
            self.__recg.endRecg = endRecg

    def setEP(self, endParse, mRecgStatus, mScore):
        if (None == self.__recg.endParse):
            self.__recg.endParse = endParse
            self.__recg.recgStatus = mRecgStatus
            self.__recg.score = mScore
        elif (mScore > self.__recg.score):
            self.__recg.endParse = endParse
            self.__recg.recgStatus = mRecgStatus
            self.__recg.score = mScore

    def getEndStatus(self):
        return self.__endStatus

    def getStartTime(self):
        return self.__startTime

    def getEndTime(self):
        return self.__endTime

    def getRecogResult(self):
        return self.__recg

    def printResult(self):
        if (None != self.__startTime and None != self.__endTime):
            self.totalCost = self.__endTime - self.__startTime
            print("totalCost:\t\t" + str(self.totalCost))
        else:
            print("loadCost:\t\tnull")
        if (None != self.frameID):
            print("frameID:\t\t" + str(self.frameID))
        else:
            print("frameID:\t\tnull")
        if (None != self.caseID):
            print("caseID:\t\t\t" + self.caseID)
        else:
            print("caseID:\t\t\tnull")
        if (None != self.__startTime):
            print("startTime:\t\t" + str(self.__startTime))
        else:
            print("startTime:\t\tnull")
        if (None != self.__endTime):
            print("endTime:\t\t" + str(self.__endTime))
        else:
            print("endTime:\t\tnull")
        if (None != self.__endStatus):
            print("endStatus:\t\t" + self.__endStatus)
        else:
            print("endStatus:\t\tnull")

        if (None != self.__startLoad):
            print("startLoadTime:\t" + str(self.__startLoad))
        else:
            print("startLoadTime:\tnull")
        if (None != self.__endLoad):
            print("endLoadTime:\t" + str(self.__endLoad))
        else:
            print("endLoadTime:\tnull")
        if (None != self.__startLoad and None != self.__endLoad):
            self.loadCost = self.__endLoad - self.__startLoad
            print("loadCost:\t\t" + str(self.loadCost))
        else:
            print("loadCost:\t\tnull")
        if (None != self.loadFile):
            print("loadFile:\t\t" + self.loadFile)
        else:
            print("loadFile:\t\tnull")

        if (None != self.__startDet):
            print("startDetTime:\t" + str(self.__startDet))
        else:
            print("startDetTime:\tnull")
        if (None != self.__endDet):
            print("endDetTime:\t\t" + str(self.__endDet))
        else:
            print("endDetTime:\t\tnull")
        if (None != self.__startDet and None != self.__endDet):
            self.detCost = self.__endDet - self.__startDet
            print("detCost:\t\t" + str(self.detCost))
        else:
            print("detCost:\t\tnull")
        if (None != self.detNum):
            print("detNum:\t\t\t" + str(self.detNum))
        else:
            print("detNum:\t\t\tnull")
        self.__recg.printResult()
        print(os.linesep)

class CaseInfo:
    def __init__(self, name, mQuitLastTime, mRemainInRecgQueue, \
        mIsLastRecgBack, iFaceOccurTime):
        self.objName = name
        self.quitLastTime = mQuitLastTime
        self.remainInRecgQueue = mRemainInRecgQueue
        self.isLastRecgBack = mIsLastRecgBack
        self.firstFaceOccurTime = iFaceOccurTime
        self.firstDetTime = None
        self.firstParseTime = None

        self.skipFrames = 0
        self.allFrames = 0
        self.detFrames = 0
        self.recgFrames = 0
        self.recgValidFrames = 0

        self.detAvgCost = 0
        self.recgAvgCost = 0
        self.recgValidAvgCost = 0
        self.recgInvalidAvgCost = 0
        self.parseAvgCost = 0
        self.loadAvgCost = 0

        self.detMaxCost = 0
        self.recgMaxCost = 0
        self.recgValidMaxCost = 0
        self.recgInvalidMaxCost = 0
        self.parseMaxCost = 0
        self.loadMaxCost = 0

        self.recgAllAvgTime = {}
        self.recgAvgScore = {}
        self.recgIDs = {}
        self.recgSuccess = 0
        self.recgCorrectTime = []

    def setIsLastRecgBack(self, mIsLastRecgBack):
        self.isLastRecgBack = mIsLastRecgBack

    def printResult(self, mWorkSheet, mCase):
        if (None == self.objName):
            return
        mWorkSheet.title = self.objName
        writeExcel = []
        writeExcel0Add = []
        writeExcel.append(mCase)
        print("objName: " + self.objName)

        writeExcel.append(str(self.skipFrames + self.allFrames))
        print("allFrames: " + str(self.skipFrames + self.allFrames))

        writeExcel.append(str(self.allFrames))
        print("allInputFrames: " + str(self.allFrames))

        writeExcel.append(str(self.detFrames))
        print("detFrames: " + str(self.detFrames))

        writeExcel.append(str(self.recgFrames))
        print("recgNum: " + str(self.recgFrames))

        writeExcel.append(str(self.recgValidFrames))
        print("recgValidNum: " + str(self.recgValidFrames))

        if 0 == self.recgFrames:
            tmpI = "NA"
        else:
            tmpI = round(self.recgValidFrames / float(self.recgFrames), 1)
        writeExcel.append(str(tmpI))
        print("recgValidRatio: " + str(tmpI))

        writeExcel.append(str(self.loadMaxCost))
        print("loadMaxCost: " + str(self.loadMaxCost))

        writeExcel.append(str(self.detMaxCost))
        print("detMaxCost: " + str(self.detMaxCost))

        if (0 == self.recgMaxCost):
            tmpI = "NA"
        else:
            tmpI = self.recgMaxCost
        writeExcel.append(str(tmpI))
        print("recgMaxCost: " + str(tmpI))

        if (0 == self.recgValidMaxCost):
            tmpI = "NA"
        else:
            tmpI = self.recgValidMaxCost
        writeExcel.append(str(tmpI))
        print("recgValidMaxCost: " + str(tmpI))

        if (0 == self.recgInvalidMaxCost):
            tmpI = "NA"
        else:
            tmpI = self.recgInvalidMaxCost
        writeExcel.append(str(tmpI))
        print("recgInvalidMaxCost: " + str(tmpI))

        if (0 == self.parseMaxCost):
            writeExcel.append("NA")
            print("parseMaxCost: NA")
        else:
            writeExcel.append(str(self.parseMaxCost))
            print("parseMaxCost: " + str(self.parseMaxCost))

        if "NA" != self.loadAvgCost and 0 != self.loadAvgCost:
            tmpI = int(self.loadAvgCost)
        else:
            tmpI = "NA"
        writeExcel.append(str(tmpI))
        print("loadAvgCost: " + str(tmpI))

        if "NA" != self.detAvgCost and 0 != self.detAvgCost:
            tmpI = int(self.detAvgCost)
        else:
            tmpI = "NA"
        writeExcel.append(str(tmpI))
        print("detAvgCost: " + str(tmpI))

        if "NA" != self.recgAvgCost and 0 != self.recgAvgCost:
            tmpI = int(self.recgAvgCost)
        else:
            tmpI = "NA"
        writeExcel.append(str(tmpI))
        print("recgAvgCost(contain recg failed cases): " + str(tmpI))

        if "NA" != self.recgValidAvgCost and 0 != self.recgValidAvgCost:
            tmpI = int(self.recgValidAvgCost)
        else:
            tmpI = "NA"
        writeExcel.append(str(tmpI))
        print("recgValidAvgCost: " + str(tmpI))

        if "NA" != self.recgInvalidAvgCost and 0 != self.recgInvalidAvgCost:
            tmpI = int(self.recgInvalidAvgCost)
        else:
            tmpI = "NA"
        writeExcel.append(str(tmpI))
        print("recgInvalidAvgCost: " + str(tmpI))

        if (0 == self.parseAvgCost) or ("NA" == self.parseAvgCost):
            writeExcel.append("NA")
            print("parseAvgCost: NA")
        else:
            writeExcel.append(str(int(self.parseAvgCost)))
            print("parseAvgCost: " + str(round(self.parseAvgCost, 0)))

        # firstDetTime firstFaceOccurTime
        if (None != self.firstFaceOccurTime and None != self.firstParseTime):
            tmp0I = self.firstParseTime - self.firstFaceOccurTime
            writeExcel.append(str(tmp0I))
            print("RecgFeelCost: " + str(tmp0I))
        else:
            writeExcel.append("NA")
            print("RecgFeelCost: NA")

        if (None != self.firstFaceOccurTime and None != self.firstDetTime):
            tmp0I = self.firstDetTime - self.firstFaceOccurTime
            writeExcel.append(str(tmp0I))
            print("DetFeelCost: " + str(tmp0I))
        else:
            writeExcel.append("NA")
            print("DetFeelCost: NA")

        if (None != self.firstParseTime and None != self.firstDetTime):
            tmp0I = self.firstParseTime - self.firstDetTime
            writeExcel.append(str(tmp0I))
            print("FeelCost: " + str(tmp0I))
        else:
            writeExcel.append("NA")
            print("FeelCost: NA")

        for it in self.recgIDs.keys():
            if (-1 != it.find(self.objName)):
                self.recgSuccess = self.recgIDs.get(it)
            tmp0F = self.recgAvgScore.get(it) / float(self.recgIDs.get(it))
            tmp0F = int(tmp0F)
            tmp1F = self.recgAllAvgTime.get(it) / float(self.recgIDs.get(it))
            tmp1F = int(tmp1F)
            print(it + " num: " + str(self.recgIDs.get(it)) \
                + "\naverage score: " + str(tmp0F) \
                + "\naverage recgAllTime: " + str(tmp1F))
            writeExcel0Add.append(it)
            writeExcel0Add.append(str(self.recgIDs.get(it)))
            writeExcel0Add.append(str(tmp0F))
            writeExcel0Add.append(str(tmp1F))

        if (0 == self.recgValidFrames):
            writeExcel.append("NA")
            print("recgSuccessRate: NA")
        else:
            tmp0F = round(self.recgSuccess / float(self.recgValidFrames), 2)
            writeExcel.append(str(tmp0F))
            print("recgSuccessRate: " + str(tmp0F))

        writeExcel.append(str(self.quitLastTime))
        print("quitLastTime: " + str(self.quitLastTime))
        writeExcel.append(str(self.remainInRecgQueue))
        print("remainInRecgQueue: " + str(self.remainInRecgQueue))
        writeExcel.append(str(self.isLastRecgBack))
        print("isLastRecgBack: " + str(self.isLastRecgBack))

        ifs = 0
        iss = 0
        its = 0
        tmpImin = sys.maxsize
        tmpIminFile = None
        tmpIminSN = None
        for it in self.recgCorrectTime:
            tmp0I = it[0] - self.firstFaceOccurTime
            if tmp0I < tmpImin:
                tmpImin = tmp0I
                tmpIminFile = it[1]
                tmpIminSN = it[2]

            if tmp0I < 2000:
                ifs += 1
            elif tmp0I < 5000:
                iss += 1
            elif tmp0I < 15000:
                its += 1

        if (sys.maxsize == tmpImin):
            tmpImin = "NA"
        if (None == tmpIminFile):
            tmpIminFile = "NA"
        if (None == tmpIminSN):
            tmpIminSN = "NA"

        writeExcel.append(str(ifs))
        writeExcel.append(str(iss + ifs))
        writeExcel.append(str(its + iss + ifs))
        writeExcel.append(str(tmpImin))
        # writeExcel.append(str(len(self.recgCorrectTime)))
        writeExcel.append(str(tmpIminFile))
        writeExcel.append(str(tmpIminSN))

        ColNum = None
        for letter in range(2, 100):
            ColNum = get_column_letter(letter)
            if (None == mWorkSheet[ColNum + '1'].value):
                break
        for idx in range(1, len(writeExcel) + 1):
            mWorkSheet[ColNum + str(idx)] = writeExcel[idx - 1]
        startIdx = len(writeExcel) + 2
        endIdx = startIdx + len(writeExcel0Add)
        for idx in range(startIdx, endIdx, 4):
            if (-1 != writeExcel0Add[idx - startIdx].find("NA")):
                mWorkSheet[ColNum + str(idx)] = writeExcel0Add[idx - startIdx] \
                + "-invalid"
            elif (-1 != writeExcel0Add[idx - startIdx].find(self.objName)):
                mWorkSheet[ColNum + str(idx)] = writeExcel0Add[idx - startIdx] \
                + "-correct"
            else:
                mWorkSheet[ColNum + str(idx)] = writeExcel0Add[idx - startIdx] \
                + "-error"
            mWorkSheet['A' + str(idx)] = "识别结果"
            idx += 1
            mWorkSheet[ColNum + str(idx)] = writeExcel0Add[idx - startIdx]
            mWorkSheet['A' + str(idx)] = "次数"
            idx += 1
            mWorkSheet[ColNum + str(idx)] = writeExcel0Add[idx - startIdx]
            mWorkSheet['A' + str(idx)] = "平均分数"
            idx += 1
            mWorkSheet[ColNum + str(idx)] = writeExcel0Add[idx - startIdx]
            mWorkSheet['A' + str(idx)] = "平均识别全程耗时(ms)"

def printDict(myDict):
    for it in myDict.keys():
        obj = myDict.get(it)
        obj.printResult()

def getCaseDict(myDict, mNameCaseDict, mOriList, iCaseFfoTsDict):
    s = {}
    for it in myDict.keys():
        name = None
        cID = myDict.get(it).caseID
        if cID not in s.keys():
            for x in mNameCaseDict.keys():
                if cID in mNameCaseDict.get(x):
                    name = x
                if None != name:
                    break

            tmp0S = None
            tmp1S = None
            tmp2S = None
            tmp3S = None
            tmp0I = None
            tmp1I = 0
            tmpImin = None

            print(name)
            print(it)
            print(cID)

            for key in iCaseFfoTsDict.keys():
                if cID in key:
                    tmp0S = cID + "/" + iCaseFfoTsDict[key]

            for line in mOriList:
                if ((-1 != line.find("quitCost: ")) and (-1 != line.find(cID))):
                    tmpSplits = line.split(" arrayRemain: ")
                    tmp1S = tmpSplits[0].split("quitCost: ")[1]
                    tmp2S = tmpSplits[1]

                if ((-1 != line.find("Running remains: ")) and (-1 != line.find(cID))):
                    if tmp1I < int(line.split("frameID: ")[1].split(" CaseIndex: ")[0]):
                        tmp1I = int(line.split("frameID: ")[1].split(" CaseIndex: ")[0])
                        tmp3S = line.split("Running remains: ")[1]

                if -1 != line.find("/" + cID + "/"):
                    if None == tmpImin:
                        tmpImin = int(line.split("|D||")[0])
                    else:
                        tmpImin = min(tmpImin, int(line.split("|D||")[0]))

                if -1 != line.find(tmp0S):
                    tmp0I = int(line.split("|D||")[0])

            if None == tmp0I:
                tmp0I = tmpImin

            if None == tmp3S:
                print("Old log~")
            elif int(tmp3S) < int(tmp2S):
                tmp2S = tmp3S

            print(tmp0I)
            s[cID] = CaseInfo(name, int(tmp1S), int(tmp2S), True, tmp0I)
        else:
            if "recoStart" == myDict.get(it).getEndStatus():
                s[cID].setIsLastRecgBack(False)
    return s

def writeCaseMap(mCaseDict, mInfoDict, mList, mWorkSheet):
    for x in mCaseDict.keys():
        objX = mCaseDict.get(x)
        for y in mList:
            if (-1 != y.find("status: loadingMiss")) and (-1 != y.find(x)):
                objX.skipFrames += 1
    for x in mCaseDict.keys():
        print("CaseID: " + x)
        detN = 0
        recgN = 0
        parseN = 0
        loadN = 0
        recgVN = 0
        recgIVN = 0
        objX = mCaseDict.get(x)
        for y in mInfoDict.keys():
            objY = mInfoDict.get(y)
            if (x == objY.caseID):
                objZ = objY.getRecogResult()
                objX.allFrames += 1
                if (0 < objY.detNum):
                    objX.detFrames += 1
                    if (None == objX.firstDetTime):
                        objX.firstDetTime = objY.getStartTime()
                    elif (objX.firstDetTime > objY.getStartTime()):
                        objX.firstDetTime = objY.getStartTime()
                if (None != objZ.recgStatus):
                    tmp = objZ.recgStatus
                    objX.recgFrames += 1
                    if ("NA" != tmp):
                        objX.recgValidFrames += 1
                        if (None == objX.firstParseTime):
                            objX.firstParseTime = objY.getEndTime()
                        elif (objX.firstParseTime > objY.getEndTime()):
                            objX.firstParseTime = objY.getEndTime()
                    if (tmp in objX.recgIDs):
                        objX.recgIDs[tmp] = objX.recgIDs.get(tmp) + 1
                        objX.recgAvgScore[tmp] = objX.recgAvgScore.get(tmp) \
                        + objZ.score
                        objX.recgAllAvgTime[tmp] = objX.recgAllAvgTime.get(tmp) \
                        + objY.totalCost
                    else:
                        objX.recgIDs[tmp] = 1
                        objX.recgAvgScore[tmp] = objZ.score
                        objX.recgAllAvgTime[tmp] = objY.totalCost

                    if (-1 != tmp.find(objX.objName)):
                        objX.recgCorrectTime.append((int(objZ.endParse), \
                            objY.fileID, objY.fileSN))

                if (None != objY.loadCost):
                    loadN += 1
                    objX.loadAvgCost += objY.loadCost
                    if (objX.loadMaxCost < objY.loadCost):
                        objX.loadMaxCost = objY.loadCost
                if (None != objY.detCost):
                    detN += 1
                    objX.detAvgCost += objY.detCost
                    if (objX.detMaxCost < objY.detCost):
                        objX.detMaxCost = objY.detCost
                if (None != objY.getRecogResult().recgCost):
                    recgN += 1
                    objX.recgAvgCost += objY.getRecogResult().recgCost
                    if (objX.recgMaxCost < objY.getRecogResult().recgCost):
                        objX.recgMaxCost = objY.getRecogResult().recgCost
                    if (objY.getRecogResult().recgIsValid):
                        recgVN += 1
                        objX.recgValidAvgCost += objY.getRecogResult().recgCost
                        if (objX.recgValidMaxCost < objY.getRecogResult().recgCost):
                            objX.recgValidMaxCost = objY.getRecogResult().recgCost
                    else:
                        recgIVN += 1
                        objX.recgInvalidAvgCost += objY.getRecogResult().recgCost
                        if (objX.recgInvalidMaxCost < objY.getRecogResult().recgCost):
                            objX.recgInvalidMaxCost = objY.getRecogResult().recgCost

                if (None != objY.getRecogResult().parseCost):
                    parseN += 1
                    objX.parseAvgCost += objY.getRecogResult().parseCost
                    if (objX.parseMaxCost < objY.getRecogResult().parseCost):
                        objX.parseMaxCost = objY.getRecogResult().parseCost
        if (0 != loadN):
            objX.loadAvgCost /= float(loadN)
        else:
            objX.loadAvgCost = "NA"
        if (0 != detN):
            objX.detAvgCost /= float(detN)
        else:
            objX.detAvgCost = "NA"
        if (0 != recgN):
            objX.recgAvgCost /= float(recgN)
        else:
            objX.recgAvgCost = "NA"
        if (0 != recgVN):
            objX.recgValidAvgCost /= float(recgVN)
        else:
            objX.recgValidAvgCost = "NA"
        if (0 != recgIVN):
            objX.recgInvalidAvgCost /= float(recgIVN)
        else:
            objX.recgInvalidAvgCost = "NA"
        if (0 != parseN):
            objX.parseAvgCost /= float(parseN)
        else:
            objX.parseAvgCost = "NA"
        objX.printResult(mWorkSheet, x)
        print(os.linesep)

def concurrentGroupJob(iObjMarks, iOriList, iStr):
    resList = []
    for (iObj, iMark) in iObjMarks:
        for mLine in iOriList:
            if (-1 != mLine.find("frameID: " + str(iMark) + ' ')):
                if (-1 != mLine.find("recoBad")):
                    epTime = int(mLine.split("|D||")[0])
                    rs = "NA"
                    confidence = float(-1)
                    iObj.setEP(epTime, rs, confidence)
                    iObj.setET(epTime, "recoBad")
                    iObj.setRecgValid(False)
                elif (-1 != mLine.find("recoParse")):
                    epTime = int(mLine.split("|D||")[0])
                    rs = mLine.split(" ID: ")[1].split(" score:")[0]
                    confidence = float(mLine.split(" score:")[1])
                    iObj.setEP(epTime, rs, confidence)
                    iObj.setET(epTime, "recoParse")
                    iObj.setRecgValid(True)
                elif (-1 != mLine.find("status: loadingStart")):
                    stTime = int(mLine.split("|D||")[0])
                    iObj.setST(stTime)
                    iObj.setET(stTime, "loadingStart")
                elif (-1 != mLine.find("status: detectStart")):
                    sdTime = int(mLine.split("|D||")[0])
                    iObj.setSD(sdTime)
                    iObj.setET(sdTime, "detectStart")
                elif (-1 != mLine.find("status: detectEnd")):
                    edTime = int(mLine.split("|D||")[0])
                    facesCnt = int(mLine.split("faceNum: ")[1])
                    iObj.setED(edTime, facesCnt)
                    iObj.setET(edTime, "detectEnd")
                elif (-1 != mLine.find("status: recoStart")):
                    srTime = int(mLine.split("|D||")[0])
                    iObj.setSR(srTime)
                    iObj.setET(srTime, "recoStart")
                elif (-1 != mLine.find("status: recoEnd")):
                    erTime = int(mLine.split("|D||")[0])
                    iObj.setER(erTime)
                    iObj.setET(erTime, "recoEnd")
        resList.append((iMark, iObj))
    sys.stdout.write(iStr)
    sys.stdout.flush()
    return resList

def concurrentJob(iObj, iOriList, iMark, iStr):
    # print(iObj.frameID)
    for mLine in iOriList:
        if (-1 != mLine.find("frameID: " + str(iMark) + ' ')):
            if (-1 != mLine.find("recoBad")):
                epTime = int(mLine.split("|D||")[0])
                rs = "NA"
                confidence = float(-1)
                iObj.setEP(epTime, rs, confidence)
                iObj.setET(epTime, "recoBad")
                iObj.setRecgValid(False)
            elif (-1 != mLine.find("recoParse")):
                epTime = int(mLine.split("|D||")[0])
                rs = mLine.split(" ID: ")[1].split(" score:")[0]
                confidence = float(mLine.split(" score:")[1])
                iObj.setEP(epTime, rs, confidence)
                iObj.setET(epTime, "recoParse")
                iObj.setRecgValid(True)
            elif (-1 != mLine.find("status: loadingStart")):
                stTime = int(mLine.split("|D||")[0])
                iObj.setST(stTime)
                iObj.setET(stTime, "loadingStart")
            elif (-1 != mLine.find("status: detectStart")):
                sdTime = int(mLine.split("|D||")[0])
                iObj.setSD(sdTime)
                iObj.setET(sdTime, "detectStart")
            elif (-1 != mLine.find("status: detectEnd")):
                edTime = int(mLine.split("|D||")[0])
                facesCnt = int(mLine.split("faceNum: ")[1])
                iObj.setED(edTime, facesCnt)
                iObj.setET(edTime, "detectEnd")
            elif (-1 != mLine.find("status: recoStart")):
                srTime = int(mLine.split("|D||")[0])
                iObj.setSR(srTime)
                iObj.setET(srTime, "recoStart")
            elif (-1 != mLine.find("status: recoEnd")):
                erTime = int(mLine.split("|D||")[0])
                iObj.setER(erTime)
                iObj.setET(erTime, "recoEnd")
    sys.stdout.write(iStr)
    sys.stdout.flush()
    return iMark, iObj

### Params region
if __name__ == '__main__':
    print(sys.version)
    print(time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()))
    strStartTime = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime())
    startTime = time.time()

    if False:
        xlsFileR, xlsFileW, srcRoot, checkList = parseOpt()
    else:
        xlsFileR = "/home/devin/Desktop/G200/FrSet2018.xlsx"
        xlsFileW = "/home/devin/Desktop/G200/RoiChange/bigger2/newInterimData.xlsx"
        srcRoot = "/home/devin/Desktop/G200/RoiChange/bigger2/"
        checkList = ["zhuyawen", "xinglj", "baoyuandong", "daiyi", \
            "peiyi", "sunhaiyan"]

    futuresList = []
    executor = ProcessPoolExecutor(max_workers = 36)

    ### Job region
    nameCaseDict = {}
    caseFfoTsDict = {}   # FfoTs: First face occurrence Time stamp
    if os.path.isfile(xlsFileR):
        wb = load_workbook(filename = xlsFileR)
    else:
        print("No xls files for reading!!!")
        sys.exit(0)

    print(wb.sheetnames)
    for it in wb.sheetnames:
        nameCase = []
        sheet = wb[it]
        for x in range(2, 46):
            tmp = sheet.cell(x, 5).value
            if (None != tmp):
                nameCase.append(str(tmp))
                tmpTuple = (it, str(tmp))
                caseFfoTsDict[tmpTuple] = str(sheet.cell(x, 7).value)
        nameCaseDict[it] = nameCase

    # print("caseFfoTsDict:")
    # print(len(caseFfoTsDict.keys()))
    # for it in caseFfoTsDict.keys():
    #     if ("0830111811" in it):
    #         print(it)
    #         print(caseFfoTsDict[it])
    # sys.exit(0)

    # print("nameCaseDict:")
    # for it in nameCaseDict.keys():
    #     print(it)
    #     obj = nameCaseDict.get(it)
    #     print(obj)
    # sys.exit(0)

    if os.path.isfile(xlsFileW):
        wb = load_workbook(filename = xlsFileW)
    else:
        wb = Workbook()

    for suffix in checkList:
        ws = wb.create_sheet("newsheet", 0)
        ws["A1"] = "视频片段编号"
        ws["A2"] = "视频片段总帧数"
        ws["A3"] = "测试所用帧数"
        ws["A4"] = "检测有效帧数"
        ws["A5"] = "识别次(帧)数"
        ws["A6"] = "有效识别次(帧)数"
        ws["A7"] = "有效识别率"
        ws["A8"] = "最大单帧加载耗时(ms)"
        ws["A9"] = "最大单帧检测耗时(ms)"
        ws["A10"] = "最大识别(含无效识别)耗时(ms)"
        ws["A11"] = "最大有效识别耗时(ms)"
        ws["A12"] = "最大无效识别耗时(ms)"
        ws["A13"] = "最大解析耗时(ms)"
        ws["A14"] = "平均单帧加载耗时(ms)"
        ws["A15"] = "平均单帧检测耗时(ms)"
        ws["A16"] = "平均识别(含无效识别)耗时(ms)"
        ws["A17"] = "平均有效识别耗时(ms)"
        ws["A18"] = "平均无效识别耗时(ms)"
        ws["A19"] = "平均解析耗时(ms)"
        ws["A20"] = "主观耗时(ms):首识别-人出现 时差"
        ws["A21"] = "主观耗时(ms):首检测-人出现 时差"
        ws["A22"] = "主观耗时(ms):首识别-首检测 时差"
        ws["A23"] = "有效帧识别正确率"
        ws["A24"] = "退出等待耗时(ms)"
        ws["A25"] = "识别队列剩余"
        ws["A26"] = "识别结果全返回"
        ws["A27"] = "2秒内识别正确次数"
        ws["A28"] = "5秒内识别正确次数"
        ws["A29"] = "15秒内识别正确次数"
        ws["A30"] = "识别正确-人出现 最短耗时"
        ws["A31"] = "首正确识别测试帧"
        ws["A32"] = "首正确识别测试帧序号"

        ws.column_dimensions[get_column_letter(1)].width = 34
        for i in range(2, 40):
            ws.column_dimensions[get_column_letter(i)].width = 18

        singleRoot = srcRoot + '/' + suffix + '/'
        futuresList = []
        oriList = []
        infoDict = {}
        caseDict = {}
        if os.path.exists(singleRoot):
            for rt, dirs, files in os.walk(singleRoot):
                for name in files:
                    print("ORI: " + os.path.join(rt, name))
                    with open(os.path.join(rt, name), 'r', encoding = 'UTF-8') as f:
                        tmpList = f.readlines()
                    for line in tmpList:
                        oriList.append(line)
            for line in oriList:
                if -1 != line.find("file: "):
                    tmpStr = line.split("file: ")[1].split(" status:")[0]
                    if ("null" == tmpStr):
                        print("One case finish~")
                    else:
                        elTime = int(line.split("|D||")[0])
                        dictKey = str(line.split("frameID: ")[1].split(" file: ")[0])
                        cID = line.split("/")[3]
                        fn = line.split("/")[4].split(" status:")[0]
                        sn = int(line.split("fileSN: ")[-1])
                        tmpStr = tmpStr.split("/")[-1]
                        tmpObj = LogInfo(dictKey, tmpStr, sn)
                        tmpObj.setEL(elTime, cID, fn)
                        tmpObj.setET(elTime, "loadingEnd")
                        infoDict[dictKey] = tmpObj
        else:
            print("No Source!!!")
            sys.exit(0)

        index = 0
        groupsize = 500
        objMarkList = []
        its = []
        for it in infoDict.keys():
            index += 1
            tmpS = str(index) + '/' + str(len(infoDict.keys())) + "\n"
            obj = infoDict.get(it)

            objMarkList.append((obj, it))

            (d, m) = divmod(index, groupsize)
            if 0 == m:
                future = executor.submit(concurrentGroupJob, objMarkList, oriList, tmpS)
                futuresList.append(future)
                objMarkList = []
        future = executor.submit(concurrentGroupJob, objMarkList, oriList, tmpS)
        futuresList.append(future)

        wait(futuresList)
        for future in futuresList:
            resList = future.result()
            for tmpT in resList:
                infoDict[tmpT[0]] = tmpT[1]

        midTime = time.time()
        strMidTime = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime())
        printDict(infoDict)
        caseDict = getCaseDict(infoDict, nameCaseDict, oriList, caseFfoTsDict)
        writeCaseMap(caseDict, infoDict, oriList, ws)

        print("Lines in all files: " + str(len(oriList)))
        print("Valid frames num: " + str(len(infoDict)))

    wbws = wb.create_sheet("caseIndex", 0)
    for i in range(1, 5):
        wbws.column_dimensions[get_column_letter(i)].width = 6
    for i in range(5, 40):
        wbws.column_dimensions[get_column_letter(i)].width = 14

    wbl = load_workbook(filename = xlsFileR)
    sheet = wbl[wbl.sheetnames[0]]
    for row in range(1, 46):
        for col in range(1, 5):
            wbws.cell(row, col).value = sheet.cell(row, col).value

    for it in wbl.sheetnames:
        sheet = wbl[it]
        addEntriesCol = wbws.max_column + 1
        addEntriesRow = 1
        wbws.cell(addEntriesRow, addEntriesCol).value = it
        for thing in sheet['E2' : 'E45']:
            addEntriesRow += 1
            wbws.cell(addEntriesRow, addEntriesCol).value = thing[0].value

    wb.save(xlsFileW)

    print(os.linesep)
    strEndTime = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime())
    endTime = time.time()
    print("Start: " + strStartTime)
    print("Mid: " + strMidTime)
    print("End: " + strEndTime)
    print("Mid: {0:,.0f}".format(midTime - startTime))
    print("Total: {0:,.0f}".format(endTime - startTime))