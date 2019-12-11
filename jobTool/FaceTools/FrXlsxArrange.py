# -*- coding: utf-8 -*-
"""
Language:   python3
Goal:       Arrange Face Recognization results with xlsx
"""
import os
import sys
import time
import subprocess
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Command line option handling
import getopt
def printUsage():
    print(u"""
Usage:
-h/--help:      show this help message
-i/--xlsFileR:  interim result excel file (.xlsx) path
-o/--xlsFileW:  output excel file (.xlsx) path
-g/--hint:      only process data matching the hint such as "普", "正,1m"

E.g.:
$>python [this.script] -i ./facerecog-interim-result.xlsx -g 正,1m
  -o ./facerecog-final-result.xlsx
""")

def parseOpt():
    try:
        opts, args = getopt.getopt(sys.argv[1:], 'hi:o:g:', \
                ['help', 'xlsFileR=', 'xlsFileW=', 'hint='])
    except getopt.GetoptError as err:
        print(str(err))
        sys.exit(1)
    for o, v in opts:
        if o in ('-h', '--help'):
            printUsage()
            sys.exit(1)
        elif o in ('-i', '--xlsFileR'):
            xlsFileR = v
        elif o in ('-o', '--xlsFileW'):
            xlsFileW = v
        elif o in ('-g', '--hint'):
            hint = v.split(",")
    for o in ('xlsFileR', 'xlsFileW'):
        if not o in dir():
            print("Error: option not specified: "+o)
            printUsage()
            sys.exit(1)
    for o in (['xlsFileR']):
        v = locals()[o]
        if not os.path.exists(v):
            print("Error: path does not exist: "+v)
            printUsage()
            sys.exit(1)
    print(xlsFileR)
    print(xlsFileW)
    if 'hint' in dir():
        print(hint)
        return (xlsFileR, xlsFileW, hint)
    else:
        return (xlsFileR, xlsFileW, [])

print(sys.version)
print(time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()))

### Defs region

### Params region
if False:
    xlsFileR, xlsFileW, hint = parseOpt()
else:
    indexSheet = "caseIndex"
    xlsFileR = "/home/devin/Downloads/tmp/newInterimData.xlsx"
    xlsFileW = "/home/devin/Downloads/tmp/SimpleData.xlsx"
    hint = ["3m", "正", "普"]
    hint = []

needRowsInfo = ["视频片段编号", "检测", "识别正确", "识别有效率", \
                "首识别-首检测 时差", "首检测-人出现 时差", "响应时间(ms)", \
                "2秒内正确识别率", "5秒内正确识别率", "15秒内正确识别率", \
                "识别正确-人出现 最短耗时", "单帧检测平均耗时"]
needRows = [1, 4, 23, 7, 22, 21, 20, 27, 28, 29, 30, 15]
individuleStartRow = 35

### Job region
matchRows = []
matchCases = []
wr = load_workbook(filename = xlsFileR)
print(wr.sheetnames)
if indexSheet in wr.sheetnames:
    wrs = wr[indexSheet]
    for r in range(2, wrs.max_row + 1):
        searchInfo = []
        for c in range(1, 5):
            searchInfo.append(wrs.cell(r, c).value)
        if (set(hint) == (set(searchInfo) & set(hint))):
            matchRows.append(r)

    for r in matchRows:
        for c in range(5, wrs.max_column + 1):
            cellValue = wrs.cell(r, c).value
            if (None != cellValue):
                matchCases.append(cellValue)
else:
    print("No case index sheet!!!")
    sys.exit(0)

print()
print("MatchCases:")
print(matchCases)

if os.path.isfile(xlsFileW):
    ww = load_workbook(filename = xlsFileW)
else:
    ww = Workbook()

newSheet = ""
for it in hint:
    newSheet += it

wws = ww.create_sheet(newSheet, 0)
wws.column_dimensions[get_column_letter(1)].width = 23
for i in range(2, 40):
    wws.column_dimensions[get_column_letter(i)].width = 15

for it in wr.sheetnames:
    if indexSheet == it:
        continue
    else:
        totalCorrectCnt = None
        wrs = wr[it]
        for c in range(2, wrs.max_column + 1):
            if (wrs.cell(1, c).value) in matchCases:
                writeRow = 1
                writeCol = wws.max_column + 1
                for r in needRows:
                    wws.cell(writeRow, 1).value = needRowsInfo[writeRow - 1]
                    if (2 == writeRow) or (3 == writeRow):
                        if "NA" != wrs.cell(r, c).value and \
                            float(wrs.cell(r, c).value) > 0:
                            wws.cell(writeRow, writeCol).value = str(1)
                        else:
                            wws.cell(writeRow, writeCol).value = str(0)
                    else:
                        wws.cell(writeRow, writeCol).value = \
                        wrs.cell(r, c).value
                    writeRow += 1

                wws.cell(writeRow, writeCol).value = "NA"
                wws.cell(writeRow, 1).value = "得分"
                for r in range(individuleStartRow, wrs.max_row + 1):
                    tmpS = wrs.cell(r, c).value
                    if None != tmpS and -1 != tmpS.find("-correct"):
                        totalCorrectCnt = int(wrs.cell(r + 1, c).value)
                        wws.cell(writeRow, writeCol).value = \
                        wrs.cell(r + 2, c).value
                        wws.cell(writeRow, 1).value = "得分"
                        break
                    else:
                        wws.cell(writeRow, writeCol).value = "NA"
                        wws.cell(writeRow, 1).value = "得分"
                writeRow += 1

                wws.cell(writeRow, writeCol).value = it
                wws.cell(writeRow, 1).value = "Name"

                for r in [8, 9, 10]:
                    tmpI = int(wws.cell(r, writeCol).value)
                    if (None == totalCorrectCnt):
                        wws.cell(r, writeCol).value = "NA" 
                    else:
                        wws.cell(r, writeCol).value = \
                        str(round(tmpI / float(totalCorrectCnt), 2))

cMax = wws.max_column + 1
wws.cell(1, cMax).value = "Average"
wws.cell(1, cMax + 2).value = "NA rate"

for r in range(2, 14):
    totalF = 0.0
    cntNA = 0
    cnt = 0
    for c in range(2, cMax):
        tmp = wws.cell(r, c).value
        if "NA" == tmp:
            cntNA += 1
            continue
        else:
            cnt += 1
            totalF += float(tmp)
    if 0 == cnt:
        wws.cell(r, cMax).value = "NA"
    else:
        wws.cell(r, cMax).value = round(totalF / cnt, 2)
    wws.cell(r, cMax + 2).value = round(cntNA / (cntNA + cnt), 2)

wws.cell(1, cMax + 1).value = "Min"
for r in [5, 6, 7, 11]:
    min0I = "NA"
    for c in range(2, cMax):
        tmp = wws.cell(r, c).value
        if "NA" == tmp:
            continue
        else:
            if "NA" == min0I:
                min0I = int(tmp)
            else:
                min0I = min(int(tmp), min0I)
    wws.cell(r, cMax + 1).value = min0I

ww.save(xlsFileW)

print(os.linesep)
print(time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()))
