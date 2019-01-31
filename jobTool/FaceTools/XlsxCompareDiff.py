# -*- coding: utf-8 -*-
"""
Language:   python3
Goal:       Compare for difference
"""
import os
import sys
import time
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill

print(sys.version)
print(time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()))

### Defs region
def diffTool(iWws, iW0rs, iW1rs, iCase, iEndRow):
    maxColumn = iWws.max_column + 1

    findA = False
    findB = False
    for c in range(2, iW0rs.max_column + 1):
        if iCase == iW0rs.cell(1, c).value:
            findA = True
        else:
            pass
    for c in range(2, iW1rs.max_column + 1):
        if iCase == iW1rs.cell(1, c).value:
            findB = True
        else:
            pass
    if findA and findB:
        pass
    else:
        return

    for c in range(2, iW0rs.max_column + 1):
        if iCase == iW0rs.cell(1, c).value:
            for r in range(1, iEndRow + 1):
                iWws.cell(r, maxColumn).value = iW0rs.cell(r, c).value
            break
        else:
            continue

    for c in range(2, iW1rs.max_column + 1):
        if iCase == iW1rs.cell(1, c).value:
            for r in range(1, iEndRow + 1):
                if r not in [1, 26, 31]:
                    if ("NA" == iWws.cell(r, maxColumn).value) or \
                    ("NA" == iW1rs.cell(r, c).value):
                        iWws.cell(r, maxColumn).value += \
                        ("-" + iW1rs.cell(r, c).value)
                    else:
                        iWws.cell(r, maxColumn).value = \
                        float(iWws.cell(r, maxColumn).value) \
                        - float(iW1rs.cell(r, c).value)
                else:
                    if iWws.cell(r, maxColumn).value == iW1rs.cell(r, c).value:
                        continue
                    else:
                        iWws.cell(r, maxColumn).value = \
                        str(iWws.cell(r, maxColumn).value) \
                        + "-" + str(iW1rs.cell(r, c).value)
                        font = Font(color = "DD0000")
                        iWws.cell(r, maxColumn).font = font
                # fill = PatternFill("solid", fgColor = "AA0000")
                # iWws.cell(r, maxColumn).fill = fill
            break
        else:
            continue

def compareTool(iWws, iW0rs, iW1rs, iCase, iEndRow):
    maxColumn = iWws.max_column + 1
    for c in range(2, iW0rs.max_column + 1):
        if iCase == iW0rs.cell(1, c).value:
            for r in range(1, iEndRow + 1):
                iWws.cell(r, maxColumn).value = iW0rs.cell(r, c).value
            break
        else:
            continue

    maxColumn = iWws.max_column + 1
    for c in range(2, iW1rs.max_column + 1):
        if iCase == iW1rs.cell(1, c).value:
            for r in range(1, iEndRow + 1):
                iWws.cell(r, maxColumn).value = iW1rs.cell(r, c).value
                font = Font(color = "DD0000")
                iWws.cell(r, maxColumn).font = font
                # fill = PatternFill("solid", fgColor = "AA0000")
                # iWws.cell(r, maxColumn).fill = fill
            break
        else:
            continue

def loadAB(iA, iB, iIndexSheet, iHint):
    if os.path.isfile(iA) and os.path.isfile(iB):
        fW0r = load_workbook(filename = iA)
        fW1r = load_workbook(filename = iB)
    else:
        print("Missing xlsx!!!")
        sys.exit(0)

    fMatchCases = []

    if iIndexSheet in fW0r.sheetnames and iIndexSheet in fW1r.sheetnames:
        fWrs = fW0r[indexSheet]
        fMatchRows = []
        for r in range(2, fWrs.max_row + 1):
            searchInfo = []
            for c in range(1, 5):
                searchInfo.append(fWrs.cell(r, c).value)
            if (set(hint) == (set(searchInfo) & set(hint))):
                fMatchRows.append(r)

        for r in fMatchRows:
            for c in range(5, fWrs.max_column + 1):
                if (None != fWrs.cell(r, c).value):
                    fMatchCases.append(fWrs.cell(r, c).value)
                else:
                    pass

        fWrs = fW1r[indexSheet]
        fMatchRows = []
        for r in range(2, fWrs.max_row + 1):
            searchInfo = []
            for c in range(1, 5):
                searchInfo.append(fWrs.cell(r, c).value)
            if (set(hint) == (set(searchInfo) & set(hint))):
                fMatchRows.append(r)

        listLen = 0
        for r in fMatchRows:
            for c in range(5, fWrs.max_column + 1):
                if (None == fWrs.cell(r, c).value):
                    continue
                elif fWrs.cell(r, c).value in fMatchCases:
                    listLen += 1
                    continue
                else:
                    print("Dismatches in indexSheet!!!")
                    sys.exit(0)

        if listLen == len(fMatchCases):
            return fMatchCases, fW0r, fW1r
        else:
            print("Dismatches in indexSheet!!!")
            sys.exit(0)
    else:
        print("No case index sheet!!!")
        sys.exit(0)

def writeXlsx(iCompareCases, iA, iB, iWriteXlsx, iIndexSheet):
    compareEndRow = 32
    if os.path.isfile(iWriteXlsx):
        fWw = load_workbook(filename = iWriteXlsx)
    else:
        fWw = Workbook()

    if set(iA.sheetnames) != set(iA.sheetnames) & set(iB.sheetnames):
        print("Dismatches in personID!!!")
        sys.exit(0)
    else:
        for id in set(iA.sheetnames):
            if iIndexSheet == id:
                continue
            fW0rs = iA[id]
            fW1rs = iB[id]
            fWws = fWw.create_sheet(id, 0)
            fWws.column_dimensions[get_column_letter(1)].width = 34
            for c in range(2, 80):
                fWws.column_dimensions[get_column_letter(c)].width = 18
            for r in range(1, compareEndRow + 1):
                fWws.cell(r, 1).value = fW0rs.cell(r, 1).value

            for x in iCompareCases:
                diffTool(fWws, fW0rs, fW1rs, x, compareEndRow)
                # compareTool(fWws, fW0rs, fW1rs, x, compareEndRow)
    fWw.save(iWriteXlsx)

### Params region
xlsFile0R = "/home/devin/Desktop/TmpResults/InterimDataSkip.xlsx"
xlsFile1R = "/home/devin/Desktop/TmpResults/InterimDataFull.xlsx"
xlsFileW = "/home/devin/Desktop/TmpResults/InterimDataCompare.xlsx"
indexSheet = "caseIndex"
hint = ["1m"]

### Job region
rt = loadAB(xlsFile0R, xlsFile1R, indexSheet, hint)
print(rt[0])
rt = writeXlsx(rt[0], rt[1], rt[2], xlsFileW, indexSheet)

print(os.linesep)
print(time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()))