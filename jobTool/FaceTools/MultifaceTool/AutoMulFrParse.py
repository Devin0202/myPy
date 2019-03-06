# -*- coding: utf-8 -*-
"""
Language:
Goal:
PS:
"""
import sys
import os
import time

import re
import random
from concurrent.futures import ProcessPoolExecutor, wait
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import numpy as np

### Common utilities
def safeDirectory(fDir):
    if str == type(fDir):
        if os.path.sep == fDir[-1]:
            safeDir = fDir
        else:
            safeDir = fDir + os.path.sep
    else:
        print("Error type of input!!!")
        sys.exit(0)
    return safeDir

def makeDirs(fDir, fExistencePermitted = True):
    safeDir = fDir
    try:
        if not os.path.exists(safeDir):
            os.makedirs(safeDir)
        else:
            if not fExistencePermitted:
                print("The folder had been existed!!!")
                sys.exit(0)
            else:
                pass
    except Exception as e:
        print("Exception occured!!!")
        print(e)
        sys.exit(0)
    else:
        print("Create: " + safeDir + "    OK")
        return safeDir

### Definition region(Class, Functions, Constants)
class CaseFramePack:
    def __init__(self, fId):
        self.mFrameId = fId
        self.mStringPack = []
    def saveString(self, fString):
        self.mStringPack.append(fString)

def getFilesLines(fDir):
    i = 0
    oriList = []
    if os.path.exists(fDir):
        print(fDir)
        for rt, dirs, files in os.walk(fDir):
            for name in files:
                tmp = re.match("\d{2}(-\d{2}){3}\.txt", name)
                if None == tmp:
                    print("MissMatch:")
                    print(name)
                    print()
                else:
                    i += 1
                    with open(os.path.join(rt, name), 'r') as fr:
                        tmpList = fr.readlines()
                    for line in tmpList:
                        oriList.append(line)
        print("Match file number:")
        print(i)
        print()
    else:
        return None
    return oriList

def classifyLines(fLines):
    outDict = {}
    for line in fLines:
        if -1 != line.find(" file: null "):
            pass
        elif -1 != line.find(" file: "):
            case = line.split(" file: ")[1].split('/')[3]
            fid = line.split("|D||frameID: ")[1].split()[0]
            if (case, fid) not in outDict.keys():
                outDict[(case, fid)] = CaseFramePack(fid)
            else:
                pass
        elif -1 != line.find(" quitIndex: "):
            case = line.split(" quitIndex: ")[1].split()[0]
            fid = line.split("|D||frameID: ")[1].split()[0]
            if (case, fid) not in outDict.keys():
                outDict[(case, fid)] = CaseFramePack(fid)
            else:
                pass
        elif -1 != line.find(" status: loadingMiss Miss: "):
            case = line.split(" status: loadingMiss Miss: ")[1].split('/')[3]
            fid = line.split("|D||frameID: ")[1].split()[0]
            if (case, fid) not in outDict.keys():
                outDict[(case, fid)] = CaseFramePack(fid)
            else:
                pass
        else:
            pass

    for line in fLines:
        if "|D||frameID: " in line:
            fid = line.split("|D||frameID: ")[1].split()[0]
            for key in outDict.keys():
                if key[1] == fid:
                    outDict[key].saveString(line)
                else:
                    pass
        else:
            pass
    return outDict

class InfoSet:
    def __init__(self):
        # Single frame & Single case
        self.frameEntry = 0
        self.frameMissed = 0
        self.frameUsed = 0
        self.detectNum = 0
        self.recgNum = 0
        self.recgValidNum = 0
        self.recgIsPair = True
        self.recgInfo = {}
        self.minDet2Recg = None
        self.avgDet2Recg = None
        self.maxDet2Recg = None

        # Single case
        self.loadST = None
        self.loadED = None
        self.loadCost = None

        self.detectST = None
        self.detectED = None
        self.detectCost = None

        self.parseCostN = None
        self.parseCostNmax = None

        self.recgCostN = None
        self.recgCostNmax = None

        self.recgNumValid = 0
        self.recgCostNvalid = None
        self.recgCostNmaxValid = None

        self.recgNumInvalid = 0
        self.recgCostNinvalid = None
        self.recgCostNmaxInvalid = None

    def accumulate(self, fSingle):
        self.frameEntry += fSingle.frameEntry
        self.frameMissed += fSingle.frameMissed
        self.frameUsed += fSingle.frameUsed
        self.detectNum += fSingle.detectNum
        self.recgNum += fSingle.recgNum
        self.recgValidNum += fSingle.recgValidNum
        self.recgIsPair = self.recgIsPair and fSingle.recgIsPair

    def infoPrint(self, fAll = True, fneedXlsx = False):
        print("frameEntry: " + str(self.frameEntry))
        print("frameUsed: " + str(self.frameUsed))
        print("frameMissed: " + str(self.frameMissed))
        print("detectNum: " + str(self.detectNum))
        print("recgValidNum: " + str(self.recgValidNum))
        print("recgNum: " + str(self.recgNum))
        if 0 == self.recgNum:
            print("recgValidRatio: None")
        else:
            print("recgValidRatio: " \
                + str(round(self.recgValidNum / self.recgNum, 2)))
        print("recgIsPair: " + str(self.recgIsPair))
        print("avgDet2Recg: " + str(self.avgDet2Recg))
        print("minDet2Recg: " + str(self.minDet2Recg))
        print("maxDet2Recg: " + str(self.maxDet2Recg))

        if fAll:
            print("recgCostN: " + str(self.recgCostN))
            print("recgCostNmax: " + str(self.recgCostNmax))

            print("recgNumValid: " + str(self.recgNumValid))
            print("recgCostNvalid: " + str(self.recgCostNvalid))
            print("recgCostNmaxValid: " + str(self.recgCostNmaxValid))

            print("recgNumInvalid: " + str(self.recgNumInvalid))
            print("recgCostNinvalid: " + str(self.recgCostNinvalid))
            print("recgCostNmaxInvalid: " + str(self.recgCostNmaxInvalid))

            print("loadCost: " + str(self.loadCost))
            print("detectCost: " + str(self.detectCost))

            print("parseCostN: " + str(self.parseCostN))
            print("parseCostNmax: " + str(self.parseCostNmax))
        else:
            pass
        print()

        outList = []
        if fneedXlsx:
            outList.append(str(self.frameEntry))
            outList.append(str(self.frameUsed))
            outList.append(str(self.frameMissed))
            outList.append(str(self.detectNum))
            outList.append(str(self.recgValidNum))
            outList.append(str(self.recgNum))
            if 0 == self.recgNum:
                outList.append("None")
            else:
                outList.append(str(round(self.recgValidNum / self.recgNum, 2)))
            outList.append(str(self.recgIsPair))
        else:
            pass
        return outList

class InfoCalc:
    def __init__(self):
        self.recgInfo = {}

        self.det2RecgMax = None
        self.det2RecgMin = None
        self.det2RecgAvg = None
        self.det2RecgCnt = None

        self.loadMax = None
        self.loadAvg = None
        self.loadCnt = None

        self.detectMax = None
        self.detectAvg = None
        self.detectCnt = None

        self.recgMax = None
        self.recgAvg = None
        self.recgCnt = None

        self.recgVMax = None
        self.recgVAvg = None
        self.recgVCnt = None

        self.recgIMax = None
        self.recgIAvg = None
        self.recgICnt = None

        self.parseMax = None
        self.parseAvg = None
        self.parseCnt = None

    def record(self, fSingle):
        for tmpKey in fSingle.recgInfo:
            if tmpKey not in self.recgInfo:
                self.recgInfo[tmpKey] = fSingle.recgInfo[tmpKey]
            else:
                tmpArray0 = self.recgInfo[tmpKey]
                tmpArray1 = fSingle.recgInfo[tmpKey]
                tmpArray0[0] = tmpArray0[0] + tmpArray1[0]
                tmpArray0[1] = tmpArray0[1] + tmpArray1[1]
                self.recgInfo[tmpKey] = tmpArray0

        if None != fSingle.avgDet2Recg:
            if None == self.det2RecgAvg:
                self.det2RecgAvg = 0
                self.det2RecgCnt = 0
                self.det2RecgMax = -sys.maxsize
                self.det2RecgMin = sys.maxsize
            else:
                pass
            self.det2RecgMax = max(fSingle.maxDet2Recg, self.det2RecgMax)
            self.det2RecgMin = min(fSingle.minDet2Recg, self.det2RecgMin)
            self.det2RecgAvg += fSingle.avgDet2Recg
            self.det2RecgCnt += 1
        else:
            pass

        if None != fSingle.loadCost:
            if None == self.loadMax:
                self.loadMax = -sys.maxsize
                self.loadAvg = 0
                self.loadCnt = 0
            else:
                pass
            self.loadMax = max(fSingle.loadCost, self.loadMax)
            self.loadAvg += fSingle.loadCost
            self.loadCnt += 1
        else:
            pass

        if None != fSingle.detectCost:
            if None == self.detectMax:
                self.detectMax = -sys.maxsize
                self.detectAvg = 0
                self.detectCnt = 0
            else:
                pass
            self.detectMax = max(fSingle.detectCost, self.detectMax)
            self.detectAvg += fSingle.detectCost
            self.detectCnt += 1
        else:
            pass

        if None != fSingle.recgCostNmax:
            if None == self.recgMax:
                self.recgMax = -sys.maxsize
                self.recgAvg = 0
                self.recgCnt = 0
            else:
                pass
            self.recgMax = max(fSingle.recgCostNmax, self.recgMax)
            self.recgAvg += fSingle.recgCostN
            self.recgCnt += fSingle.recgNum
        else:
            pass

        if None != fSingle.recgCostNmaxValid:
            if None == self.recgVMax:
                self.recgVMax = -sys.maxsize
                self.recgVAvg = 0
                self.recgVCnt = 0
            else:
                pass
            self.recgVMax = max(fSingle.recgCostNmaxValid, self.recgVMax)
            self.recgVAvg += fSingle.recgCostNvalid
            self.recgVCnt += fSingle.recgNumValid
        else:
            pass

        if None != fSingle.recgCostNmaxInvalid:
            if None == self.recgIMax:
                self.recgIMax = -sys.maxsize
                self.recgIAvg = 0
                self.recgICnt = 0
            else:
                pass
            self.recgIMax = max(fSingle.recgCostNmaxInvalid, self.recgIMax)
            self.recgIAvg += fSingle.recgCostNinvalid
            self.recgICnt += fSingle.recgNumInvalid
        else:
            pass

        if None != fSingle.parseCostNmax:
            if None == self.parseMax:
                self.parseMax = -sys.maxsize
                self.parseAvg = 0
                self.parseCnt = 0
            else:
                pass
            self.parseMax = max(fSingle.parseCostNmax, self.parseMax)
            self.parseAvg += fSingle.parseCostN
            self.parseCnt += fSingle.recgNum
        else:
            pass

    def calc(self):
        for tmpKey in self.recgInfo:
            tmpArray0 = self.recgInfo[tmpKey]
            tmpArray0[1] = round(tmpArray0[1] / tmpArray0[0], 2)
            self.recgInfo[tmpKey] = tmpArray0

        if None != self.loadAvg and None != self.loadCnt and 0 != self.loadCnt:
            self.loadAvg = round(self.loadAvg / self.loadCnt, 3)
        else:
            self.loadAvg = None
        if None != self.detectAvg and None != self.detectCnt and 0 != self.detectCnt:
            self.detectAvg = round(self.detectAvg / self.detectCnt, 3)
        else:
            self.detectAvg = None
        if None != self.recgAvg and None != self.recgCnt and 0 != self.recgCnt:
            self.recgAvg = round(self.recgAvg / self.recgCnt, 3)
        else:
            self.recgAvg = None
        if None != self.recgVAvg and None != self.recgVCnt and 0 != self.recgVCnt:
            self.recgVAvg = round(self.recgVAvg / self.recgVCnt, 3)
        else:
            self.recgVAvg = None
        if None != self.recgIAvg and None != self.recgICnt and 0 != self.recgICnt:
            self.recgIAvg = round(self.recgIAvg / self.recgICnt, 3)
        else:
            self.recgIAvg = None
        if None != self.parseAvg and None != self.parseCnt and 0 != self.parseCnt:
            self.parseAvg = round(self.parseAvg / self.parseCnt, 3)
        else:
            self.parseAvg = None
        if None != self.det2RecgAvg and None != self.det2RecgCnt \
            and 0 != self.det2RecgCnt:
            self.det2RecgAvg = round(self.det2RecgAvg / self.det2RecgCnt, 3)
        else:
            self.det2RecgAvg = None

    def infoPrint(self, fneedXlsx = False):
        print("recgInfo: " + str(self.recgInfo))
        print("loadAvg: " + str(self.loadAvg))
        print("detectAvg: " + str(self.detectAvg))
        print("recgAvg: " + str(self.recgAvg))
        print("recgVAvg: " + str(self.recgVAvg))
        print("recgIAvg: " + str(self.recgIAvg))
        print("parseAvg: " + str(self.parseAvg))
        print()
        print("loadMax: " + str(self.loadMax))
        print("detectMax: " + str(self.detectMax))
        print("recgMax: " + str(self.recgMax))
        print("recgVMax: " + str(self.recgVMax))
        print("recgIMax: " + str(self.recgIMax))
        print("parseMax: " + str(self.parseMax))
        print()
        print("loadCnt: " + str(self.loadCnt))
        print("detectCnt: " + str(self.detectCnt))
        print("recgCnt: " + str(self.recgCnt))
        print("recgVCnt: " + str(self.recgVCnt))
        print("recgICnt: " + str(self.recgICnt))
        print("parseCnt: " + str(self.parseCnt))
        print()
        print("det2RecgAvg: " + str(self.det2RecgAvg))
        print("det2RecgCnt: " + str(self.det2RecgCnt))
        print("det2RecgMax: " + str(self.det2RecgMax))
        print("det2RecgMin: " + str(self.det2RecgMin))
        print()

        outList = []
        if fneedXlsx:
            outList.append(str(self.loadAvg))
            outList.append(str(self.detectAvg))
            outList.append(str(self.recgAvg))
            outList.append(str(self.recgVAvg))
            outList.append(str(self.recgIAvg))
            outList.append(str(self.parseAvg))
            outList.append(str(self.loadMax))
            outList.append(str(self.detectMax))
            outList.append(str(self.recgMax))
            outList.append(str(self.recgVMax))
            outList.append(str(self.recgIMax))
            outList.append(str(self.parseMax))
            outList.append(str(self.det2RecgCnt))
            outList.append(str(self.det2RecgAvg))
            outList.append(str(self.det2RecgMax))
            outList.append(str(self.det2RecgMin))

            outList.append('')
            for tmpKey in self.recgInfo:
                outList.append(tmpKey)
                tmpArray0 = self.recgInfo[tmpKey]
                outList.append(tmpArray0[0])
                outList.append(tmpArray0[1])
        else:
            pass
        return outList

def frameProcess(fStrings):
    earlyRecgStart = None
    recgIds = []
    recgEach = []
    recgStartNum = 0
    singleOne = InfoSet()
    singleOne.frameEntry = len(fStrings)
    for line in fStrings:
        if " file: " in line and ".nv21 " in line:
            singleOne.frameUsed = 1
        else:
            pass
        if " status: loadingMiss " in line:
            singleOne.frameMissed += 1
        else:
            pass
        if " qualified faceNum: " in line:
            singleOne.detectNum = int(line.split(" qualified faceNum: ")[-1])
        else:
            pass
        if " status: recoEnd" in line:
            singleOne.recgNum += 1
            recgIds.append(line.split(" recoID: ")[1].split()[0])
        else:
            pass
        if " status: recoStart" in line:
            recgStartNum += 1
        else:
            pass
        if " status: recoParse " in line:
            singleOne.recgValidNum += 1
        else:
            pass
        if " status: loadingStart" in line:
            singleOne.loadST = int(line.split("|D||")[0])
        else:
            pass
        if " status: loadingEnd" in line:
            singleOne.loadED = int(line.split("|D||")[0])
        else:
            pass
        if " status: detectStart" in line:
            singleOne.detectST = int(line.split("|D||")[0])
        else:
            pass
        if " status: detectEnd" in line:
            singleOne.detectED = int(line.split("|D||")[0])
            earlyRecgStart = singleOne.detectED
        else:
            pass

    if recgStartNum != singleOne.recgNum:
        singleOne.recgIsPair = False
    else:
        pass
    if None != singleOne.loadST and None != singleOne.loadED:
        singleOne.loadCost = singleOne.loadED - singleOne.loadST
    else:
        pass
    if None != singleOne.detectST and None != singleOne.detectED:
        singleOne.detectCost = singleOne.detectED - singleOne.detectST
    else:
        pass

    for i in recgIds:
        tmpS0 = "recoID: " + i
        recgValid = False
        recgST = None
        parseST = None
        parseED = None
        for line in fStrings:
            if tmpS0 in line and " status: recoStart" in line:
                recgST = int(line.split("|D||")[0])
            else:
                pass
            if tmpS0 in line and " status: recoEnd" in line:
                parseST = int(line.split("|D||")[0])
                recgEach.append(parseST - earlyRecgStart)
            else:
                pass
            if tmpS0 in line and "status: recgInvalidRoi" in line:
                parseED = int(line.split("|D||")[0])
                recgValid = False
            else:
                pass
            if tmpS0 in line and "status: recgValidRoi" in line:
                parseED = int(line.split("|D||")[0])
                recgValid = True
            else:
                pass
            if tmpS0 in line and " status: recoParse " in line:
                keyId = line.split(" ID: ")[1].split()[0]
                scoreId = float(line.split(" score: ")[1].split()[0])
                if keyId not in singleOne.recgInfo:
                    singleOne.recgInfo[keyId] = [1, scoreId]
                else:
                    tmpArray0 = singleOne.recgInfo[keyId]
                    tmpArray0[0] = tmpArray0[0] + 1
                    tmpArray0[1] = tmpArray0[1] + scoreId
                    singleOne.recgInfo[keyId] = tmpArray0
            else:
                pass
        if None != parseST and None != parseED:
            if None == singleOne.parseCostN:
                singleOne.parseCostN = 0
                singleOne.parseCostNmax = 0
            else:
                pass
            tmpI0 = parseED - parseST
            singleOne.parseCostN += tmpI0
            singleOne.parseCostNmax = max(tmpI0, singleOne.parseCostNmax)
        else:
            pass
        if None != parseST and None != recgST:
            if None == singleOne.recgCostN:
                singleOne.recgCostN = 0
                singleOne.recgCostNmax = 0
            else:
                pass
            tmpI0 = parseST - recgST
            singleOne.recgCostN += tmpI0
            singleOne.recgCostNmax = max(tmpI0, singleOne.recgCostNmax)

            if recgValid:
                if None == singleOne.recgCostNvalid:
                    singleOne.recgCostNvalid = 0
                    singleOne.recgCostNmaxValid = 0
                else:
                    pass
                singleOne.recgNumValid += 1
                singleOne.recgCostNvalid += tmpI0
                singleOne.recgCostNmaxValid = max(tmpI0, \
                    singleOne.recgCostNmaxValid)
            else:
                if None == singleOne.recgCostNinvalid:
                    singleOne.recgCostNinvalid = 0
                    singleOne.recgCostNmaxInvalid = 0
                else:
                    pass
                singleOne.recgNumInvalid += 1
                singleOne.recgCostNinvalid += tmpI0
                singleOne.recgCostNmaxInvalid = max(tmpI0, \
                    singleOne.recgCostNmaxInvalid)
        else:
            pass

    if recgEach:
        singleOne.minDet2Recg = np.min(recgEach)
        singleOne.avgDet2Recg = np.mean(recgEach)
        singleOne.maxDet2Recg = np.max(recgEach)
    else:
        pass
    return singleOne

def getExitInfo(fStringPack):
    for line in fStringPack:
        if "quitCost: " in line and "arrayRemain: " in line:
            quitCost = int(line.split("quitCost: ")[1].split()[0])
            arrayRemain = int(line.split("arrayRemain: ")[1])
        else:
            quitCost = None
            arrayRemain = None
    return quitCost, arrayRemain

def writeXlsx(fWriteCol, fCase, fWsList, fWs):
    align = Alignment(horizontal = 'left',vertical = 'center')

    fWs.cell(1, fWriteCol).alignment = align
    fWs.cell(1, fWriteCol).value = fCase
    row = 2
    for i in fWsList:
        fWs.cell(row, fWriteCol).alignment = align
        fWs.cell(row, fWriteCol).value = i
        row += 1

def caseProcess(fDict, fWs = None):
    caseSet = []
    for key in fDict.keys():
        caseSet.append(key[0])
    caseSet = set(caseSet)
    print("Case info:")
    print(caseSet)
    print()

    for case in caseSet:
        totalOne = InfoSet()
        calcOne = InfoCalc()
        oQuitCost = None
        OArrayRemain = None
        print("========Case Id: " + str(case) + "========")
        for key in fDict.keys():
            if case == key[0]:
                stringPack = fDict[key].mStringPack
                result = frameProcess(stringPack)
                totalOne.accumulate(result)
                calcOne.record(result)
                fDict[key] = result
                if None == oQuitCost and None == OArrayRemain:
                    oQuitCost, OArrayRemain = getExitInfo(stringPack)
                else:
                    pass
            else:
                pass
        calcOne.calc()

        wsList = []
        wsList.extend(totalOne.infoPrint(fAll = False, fneedXlsx = True))
        print("quitCost: " + str(oQuitCost))
        print("arrayRemain: " + str(OArrayRemain))
        wsList.append(str(oQuitCost))
        wsList.append(str(OArrayRemain))
        wsList.extend(calcOne.infoPrint(fneedXlsx = True))
        print()

        if None != fWs:
            writeCol = fWs.max_column + 1
            writeXlsx(writeCol, case, wsList, fWs)
        else:
            writeCol = None
    return fDict

def xlsxSetting(fXlsxWb, fTitle):
    ws = fXlsxWb.create_sheet(fTitle, 0)
    ws["A1"] = "视频片段编号"
    ws["A2"] = "日志总行数"
    ws["A3"] = "测试所用帧数"
    ws["A4"] = "测试略过帧数"
    ws["A5"] = "检测出置信目标数"
    ws["A6"] = "有效识别数"
    ws["A7"] = "识别数"
    ws["A8"] = "有效识别率"
    ws["A9"] = "识别结果全返回"

    ws["A10"] = "退出等待耗时(ms)"
    ws["A11"] = "识别队列剩余"

    ws["A12"] = "平均单帧加载耗时(ms)"
    ws["A13"] = "平均单帧检测耗时(ms)"
    ws["A14"] = "平均识别(含无效识别)耗时(ms)"
    ws["A15"] = "平均有效识别耗时(ms)"
    ws["A16"] = "平均无效识别耗时(ms)"
    ws["A17"] = "平均解析耗时(ms)"

    ws["A18"] = "最大单帧加载耗时(ms)"
    ws["A19"] = "最大单帧检测耗时(ms)"
    ws["A20"] = "最大识别(含无效识别)耗时(ms)"
    ws["A21"] = "最大有效识别耗时(ms)"
    ws["A22"] = "最大无效识别耗时(ms)"
    ws["A23"] = "最大解析耗时(ms)"
    ws["A24"] = "识别帧总数"
    ws["A25"] = "单帧识别平均耗时(ms)"
    ws["A26"] = "单帧识别最大耗时(ms)"
    ws["A27"] = "单帧识别最小耗时(ms)"

    ws.column_dimensions[get_column_letter(1)].width = 34
    for i in range(2, 24):
        ws.column_dimensions[get_column_letter(i)].width = 18
    return ws

def testTool(fDict):
    tmp = random.sample(fDict.keys(), 1)[0]
    tmp = ("0222144758", "603060441")
    print("========Test========")
    print("Case: " + tmp[0])
    print("FrameId: " + tmp[1])
    if tmp in fDict:
        print("Key found")
    else:
        print("Key not found")
    fDict[tmp].infoPrint()

if "__main__" == __name__:
    print(sys.version)
    timeStampFormat = "%Y-%m-%d %H:%M:%S"
    print(time.strftime(timeStampFormat, time.localtime()))
### Parameters region
    xlsFileW = "/home/devin/Desktop/G200/DoubleMan/skip0301/batfrog-out0/result/InterimData.xlsx"
    cSrcRoot = "/home/devin/Desktop/G200/DoubleMan/skip0301/batfrog-out0/result/"

    xlsFileW = "/home/devin/Downloads/tmp/batfrog-out0/result/InterimData.xlsx"
    cSrcRoot = "/home/devin/Downloads/tmp/batfrog-out0/result/"
### Job region
    if os.path.isfile(xlsFileW):
        os.remove(xlsFileW)
    else:
        pass
    wb = Workbook()

    cSrcRoot = safeDirectory(cSrcRoot)
    filesList = os.listdir(cSrcRoot)
    for suffix in filesList:
        tmpDir = cSrcRoot + suffix
        tmpDir = safeDirectory(tmpDir)
        result = getFilesLines(tmpDir)
        if None != result:
            ws = xlsxSetting(wb, suffix)
            result = classifyLines(result)
            result = caseProcess(result, ws)
            # testTool(result)
            # sys.exit(0)
        else:
            pass

    wb.save(xlsFileW)

    print(os.linesep)
    print(time.strftime(timeStampFormat, time.localtime()))