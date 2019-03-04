# -*- coding: utf-8 -*-
"""
Language:   python3
Goal:       Copy the hints-matched folders with xlsx file
PS:
"""
import sys
import os
import time

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import shutil

### Common utilities
def safeDirectory(fDir):
    if str == type(fDir):
        if os.path.sep == fDir[-1]:
            safeDir = fDir
        else:
            safeDir = fDir + os.path.sep
    else:
        print("Error type of input!!!")
        sys.exit(0)
    return safeDir

def makeDirs(fDir, fExistencePermitted = True):
    safeDir = fDir
    try:
        if not os.path.exists(safeDir):
            os.makedirs(safeDir)
        else:
            if not fExistencePermitted:
                print("The folder had been existed!!!")
                sys.exit(0)
            else:
                pass
    except Exception as e:
        print("Exception occured!!!")
        print(e)
        sys.exit(0)
    else:
        print("Create: " + safeDir + "    OK")
        return safeDir

### Definition region(Class, Functions, Constants)
def ExtractInfo(fXlsx, fHint = []):
    if not os.path.isfile(fXlsx):
        pass
    else:
        totalMatchCases = []
        totalRemainCases = []
        wr = load_workbook(filename = fXlsx)
        for indexSheet in wr.sheetnames:
            wrs = wr[indexSheet]
            matchRows = []
            matchCases = []
            remainRows = []
            remainCases = []
            for r in range(2, wrs.max_row + 1):
                searchInfo = []
                for c in range(1, 5):
                    searchInfo.append(wrs.cell(r, c).value)
                if (set(fHint) == (set(searchInfo) & set(fHint))):
                    matchRows.append(r)
                else:
                    remainRows.append(r)

            for r in matchRows:
                cellValue = wrs.cell(r, 5).value
                if (None != cellValue):
                    matchCases.append(cellValue)
            for r in remainRows:
                cellValue = wrs.cell(r, 5).value
                if (None != cellValue):
                    remainCases.append(cellValue)
            totalMatchCases.extend(matchCases)
            totalRemainCases.extend(remainCases)

        # print()
        # print("totalMatchCases:")
        # print(totalMatchCases)
        # print()
        # print("totalRemainCases:")
        # print(totalRemainCases)
    return totalMatchCases, totalRemainCases

if "__main__" == __name__:
    print(sys.version)
    timeStampFormat = "%Y-%m-%d %H:%M:%S"
    print(time.strftime(timeStampFormat, time.localtime()))
### Parameters region
    cPath = "/media/devin/OpenImage600/faces/"
    cNewPath00 = "/media/devin/OpenImage600/3m普正右/"
    cNewPath01 = "/media/devin/OpenImage600/1m普正左/"
    xlsFileR = "/media/devin/OpenImage600/faces/FrSet2018.xlsx"
    cHint = ["3m", "普", "正", "右"]
### Job region
    tm, tr = ExtractInfo(xlsFileR, cHint)
    print("Intersection:")
    print(set(tm) & set(tr))
    for rt, dirs, files in os.walk(cPath):
        for i in dirs:
            if i in tm:
                src = os.path.join(rt, i)
                dst = src.replace(cPath, cNewPath00)
                shutil.copytree(src, dst)
            #     print(dst)
            # if i in tr:
            #     src = os.path.join(rt, i)
            #     dst = src.replace(cPath, cNewPath01)
            #     shutil.copytree(src, dst)
                # print(dst)

    print(os.linesep)
    print(time.strftime(timeStampFormat, time.localtime()))